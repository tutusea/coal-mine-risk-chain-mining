# -*- coding: utf-8 -*-
"""
Unified pipeline entry: runs src stages 01 through 07 in order.
From project root: python run_all.py
"""

from __future__ import annotations

import importlib
import json
import sys
import traceback
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import config

# Stage module names must match src filenames (load with importlib for numeric prefixes).
PIPELINE_STAGES = [
    "01_clean_data",
    "02_build_graph",
    "03_compute_network_metrics",
    "04_mine_risk_chains",
    "05_association_rules",
    "06_make_figures",
    "07_generate_report",
]


def _ensure_output_dirs() -> None:
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    config.FIGURES_DIR.mkdir(parents=True, exist_ok=True)
    config.DATA_AUDIT_DIR.mkdir(parents=True, exist_ok=True)


def _scan_outputs_for_legacy_haz_clean(out_dir: Path) -> list[str]:
    """
    扫描正式结果目录根层 csv / xlsx / md 是否仍含字符串 Haz_clean。
    RUN_LOG.md 不在此目录，不会被扫描。
    """
    needle = "Haz_clean"
    hits: list[str] = []
    if not out_dir.is_dir():
        return hits
    for pattern in ("*.csv", "*.md", "*.xlsx"):
        for path in sorted(out_dir.glob(pattern)):
            try:
                suf = path.suffix.lower()
                if suf in {".csv", ".md"}:
                    txt = path.read_text(encoding="utf-8", errors="ignore")
                    if needle not in txt:
                        continue
                    for i, line in enumerate(txt.splitlines(), 1):
                        if needle in line:
                            hits.append(f"{path.name}: line {i} contains {needle!r}")
                            break
                elif suf == ".xlsx":
                    from openpyxl import load_workbook

                    wb = load_workbook(path, read_only=True, data_only=True)
                    try:
                        hit_this_file = False
                        for sn in wb.sheetnames:
                            ws = wb[sn]
                            for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
                                for ci, cell in enumerate(row):
                                    if cell is not None and needle in str(cell):
                                        hits.append(
                                            f"{path.name}: sheet={sn!r} row={ri} col_index={ci} contains {needle!r}"
                                        )
                                        hit_this_file = True
                                        break
                                if hit_this_file:
                                    break
                            if hit_this_file:
                                break
                    finally:
                        wb.close()
            except Exception as exc:
                hits.append(f"{path.name}: scan error ({type(exc).__name__}: {exc})")
    return hits


def _append_quality_revision_run_log(*, haz_clean_hits: list[str]) -> None:
    """在 RUN_LOG.md 末尾追加「最终结果表清洗与稳健规则输出」摘要（USE_REAL_ONLY 模式）。"""
    log_path = ROOT / "RUN_LOG.md"
    out_dir = config.OUTPUT_DIR
    ts = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M %z")

    s01: dict = {}
    p01 = out_dir / "pipeline_stage01_stats.json"
    if p01.is_file():
        try:
            s01 = json.loads(p01.read_text(encoding="utf-8"))
        except Exception:
            s01 = {}

    n_real = int(s01.get("n_rows_exported", 0))
    n_hm_nonempty = int(s01.get("n_hazard_mode_nonempty", 0))
    n_hm_uq = int(s01.get("n_hazard_mode_unique", 0))
    n_if_nonempty = int(s01.get("n_injury_form_nonempty", 0))
    n_if_uq = int(s01.get("n_injury_form_unique", 0))
    n_hs_nonempty = int(s01.get("n_hazard_source_nonempty", 0))
    n_hs_uq = int(s01.get("n_hazard_source_unique", 0))
    n_unc = int(s01.get("n_haz_uncertain", 0))

    def _xlsx_nonempty(p: Path) -> bool:
        if not p.is_file():
            return False
        try:
            return len(pd.read_excel(p, engine="openpyxl")) > 0
        except Exception:
            return False

    def _nrules(name: str) -> int:
        p = out_dir / name
        if not p.is_file():
            return 0
        try:
            return max(0, len(pd.read_csv(p, encoding="utf-8-sig")))
        except Exception:
            return 0

    n_prev_sev = _nrules("association_rules_severe_prevention.csv")
    n_prev_sev_rob = _nrules("association_rules_severe_prevention_robust.csv")
    n_cons_sev = _nrules("association_rules_severe_consequence.csv")
    n_cons_sev_rob = _nrules("association_rules_severe_consequence_robust.csv")
    n_prev_ser = _nrules("association_rules_serious_prevention.csv")

    haz_clean_flag = "否" if not haz_clean_hits else "是（见本块下列明细）"
    dedup_sev_ok = _xlsx_nonempty(out_dir / "top_30_local_severe_chains_dedup.xlsx")
    dedup_ser_ok = _xlsx_nonempty(out_dir / "top_30_local_serious_chains_dedup.xlsx")

    files = [
        str(out_dir / "hazard_value_audit.xlsx"),
        str(out_dir / "cleaned_data.xlsx"),
        str(config.DATA_AUDIT_DIR / "hazard_uncertain_values.csv"),
        str(out_dir / "local_chain_summary.xlsx"),
        str(out_dir / "full_chain_summary.xlsx"),
        str(out_dir / "top_30_local_severe_chains.xlsx"),
        str(out_dir / "top_30_local_serious_chains.xlsx"),
        str(out_dir / "top_30_local_severe_chains_dedup.xlsx"),
        str(out_dir / "top_30_local_serious_chains_dedup.xlsx"),
        str(out_dir / "association_rules_severe_prevention.csv"),
        str(out_dir / "association_rules_severe_prevention_robust.csv"),
        str(out_dir / "association_rules_severe_consequence.csv"),
        str(out_dir / "association_rules_severe_consequence_robust.csv"),
        str(out_dir / "association_rules_serious_prevention.csv"),
        str(config.FIGURES_DIR / "node_risk_ranking_prevention.png"),
        str(config.FIGURES_DIR / "node_risk_ranking_prevention.pdf"),
        str(config.FIGURES_DIR / "node_risk_ranking_consequence.png"),
        str(config.FIGURES_DIR / "node_risk_ranking_consequence.pdf"),
        str(config.FIGURES_DIR / "network_overview_local_chain.png"),
        str(config.FIGURES_DIR / "network_overview_local_chain.pdf"),
        str(config.FIGURES_DIR / "network_overview_local_chain_simplified.png"),
        str(config.FIGURES_DIR / "network_overview_local_chain_simplified.pdf"),
        str(config.FIGURES_DIR / "association_rules_severe_prevention_bar.png"),
        str(config.FIGURES_DIR / "association_rules_severe_prevention_bar.pdf"),
        str(config.FIGURES_DIR / "top_risk_chains_sankey.html"),
        str(out_dir / "report_summary.md"),
    ]

    lines = [
        f"### AUTO — 最终结果表清洗与稳健规则输出 / {ts}",
        "",
        f"- **真实样本数（cleaned_data 行数）**：{n_real}",
        f"- **HazardMode 非 Unknown 条数**：{n_hm_nonempty}；**唯一值数（非 Unknown）**：{n_hm_uq}",
        f"- **InjuryForm 非 Unknown 条数**：{n_if_nonempty}；**唯一值数（非 Unknown）**：{n_if_uq}",
        f"- **HazardSource 有效条数（非 Unknown/NotAvailable）**：{n_hs_nonempty}；**唯一值数**：{n_hs_uq}",
        f"- **uncertain（无法归类）条数**：{n_unc}",
        f"- **是否仍存在 Haz_clean（正式 outputs 根层 csv/xlsx/md）**：{haz_clean_flag}",
    ]
    if haz_clean_hits:
        for h in haz_clean_hits[:25]:
            lines.append(f"  - {h}")
        if len(haz_clean_hits) > 25:
            lines.append(f"  - … 另有 {len(haz_clean_hits) - 25} 处命中未逐条列出")
    lines.extend(
        [
            f"- **top_30_local_severe_chains_dedup.xlsx 是否非空**：{'是' if dedup_sev_ok else '否'}",
            f"- **top_30_local_serious_chains_dedup.xlsx 是否非空**：{'是' if dedup_ser_ok else '否'}",
            f"- **severe_prevention 原始规则数量**：{n_prev_sev}",
            f"- **severe_prevention robust 规则数量**：{n_prev_sev_rob}",
            f"- **severe_consequence 原始规则数量**：{n_cons_sev}",
            f"- **severe_consequence robust 规则数量**：{n_cons_sev_rob}",
            f"- **serious_prevention 规则数量**：{n_prev_ser}",
            "- **新生成/更新的主要文件路径**：",
        ]
    )
    for fp in files:
        lines.append(f"  - `{fp}`")
    lines.extend(["", "---", ""])

    with open(log_path, "a", encoding="utf-8") as f:
        f.write("\n".join(lines))


def _append_local_chain_dedup_revision_run_log() -> None:
    """在 RUN_LOG.md 末尾追加「高风险局部链条表筛选与去重修正」记录（USE_REAL_ONLY 全链路成功后）。"""
    log_path = ROOT / "RUN_LOG.md"
    out_dir = config.OUTPUT_DIR
    ts = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M %z")

    n_real = 0
    p01 = out_dir / "pipeline_stage01_stats.json"
    if p01.is_file():
        try:
            n_real = int(json.loads(p01.read_text(encoding="utf-8")).get("n_rows_exported", 0))
        except Exception:
            n_real = 0

    min_c = int(getattr(config, "MIN_CHAIN_COUNT", 3))

    n_local_rows = 0
    p_local = out_dir / "local_chain_summary.xlsx"
    if p_local.is_file():
        try:
            n_local_rows = len(pd.read_excel(p_local, engine="openpyxl"))
        except Exception:
            n_local_rows = 0

    def _read_dedup_stats(name: str, rate_col: str) -> tuple[int, int]:
        p = out_dir / name
        if not p.is_file():
            return 0, 0
        try:
            d = pd.read_excel(p, engine="openpyxl")
        except Exception:
            return 0, 0
        n = len(d)
        if n == 0 or rate_col not in d.columns:
            return n, 0
        rv = pd.to_numeric(d[rate_col], errors="coerce").fillna(0.0)
        n_zero = int((rv <= 0).sum())
        return n, n_zero

    n_sev_dedup, n_sev_rate_le0 = _read_dedup_stats("top_30_local_severe_chains_dedup.xlsx", "severe_rate")
    n_ser_dedup, n_ser_rate_le0 = _read_dedup_stats("top_30_local_serious_chains_dedup.xlsx", "serious_rate")

    report_path = out_dir / "report_summary.md"
    report_prefers_dedup = False
    if report_path.is_file():
        try:
            txt = report_path.read_text(encoding="utf-8", errors="ignore")
            report_prefers_dedup = (
                "top_30_local_severe_chains_dedup" in txt and "5.2 高风险局部链条，按严重伤害率和频次筛选" in txt
            )
        except Exception:
            report_prefers_dedup = False

    files = [
        str(out_dir / "local_chain_summary.xlsx"),
        str(out_dir / "top_30_local_severe_chains.xlsx"),
        str(out_dir / "top_30_local_serious_chains.xlsx"),
        str(out_dir / "top_30_local_severe_chains_dedup.xlsx"),
        str(out_dir / "top_30_local_serious_chains_dedup.xlsx"),
        str(out_dir / "report_summary.md"),
        str(ROOT / "src" / "04_mine_risk_chains.py"),
        str(ROOT / "src" / "07_generate_report.py"),
    ]

    lines = [
        "### 高风险局部链条表筛选与去重修正",
        "",
        f"- **记录时间**：{ts}",
        "",
        f"1. **真实样本数（pipeline_stage01_stats.json / n_rows_exported）**：{n_real}",
        f"2. **MIN_CHAIN_COUNT 当前值**：{min_c}",
        f"3. **local_chain_summary.xlsx 总行数**：{n_local_rows}",
        f"4. **top_30_local_severe_chains_dedup.xlsx 行数**：{n_sev_dedup}",
        f"5. **top_30_local_severe_chains_dedup.xlsx 中 severe_rate≤0 的行数（应为 0）**：{n_sev_rate_le0}",
        f"6. **top_30_local_serious_chains_dedup.xlsx 行数**：{n_ser_dedup}",
        f"7. **top_30_local_serious_chains_dedup.xlsx 中 serious_rate≤0 的行数（应为 0）**：{n_ser_rate_le0}",
        f"8. **report_summary.md 是否已优先使用 dedup 并采用新 5.2 标题**：{'是' if report_prefers_dedup else '否（请检查报告生成是否成功）'}",
        "9. **本次涉及的主要文件路径**：",
    ]
    for fp in files:
        lines.append(f"   - `{fp}`")
    lines.extend(["", "---", ""])

    with open(log_path, "a", encoding="utf-8") as f:
        f.write("\n".join(lines))


def run_stage(module_name: str) -> None:
    mod = importlib.import_module(f"src.{module_name}")
    run = getattr(mod, "run", None)
    if callable(run):
        run()
    else:
        print(f"SKIP: src.{module_name} has no callable run(), skipped.", flush=True)


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Run risk-chain pipeline or optional sensitivity analysis.")
    parser.add_argument(
        "--sensitivity",
        action="store_true",
        help="Run sensitivity analysis only (stage 08); does not re-run main pipeline.",
    )
    args = parser.parse_args()

    if args.sensitivity:
        print("Stage: sensitivity analysis (08)", flush=True)
        try:
            mod = importlib.import_module("src.08_sensitivity_analysis")
            mod.run()
        except Exception as exc:
            print("", flush=True)
            print("ERROR: Sensitivity analysis failed.", flush=True)
            print(f"Exception type: {type(exc).__name__}", flush=True)
            print(f"Message: {exc}", flush=True)
            print("", flush=True)
            print("Full traceback:", flush=True)
            traceback.print_exc()
            return 1
        print("OK: Sensitivity analysis finished.", flush=True)
        return 0

    _ensure_output_dirs()

    n = len(PIPELINE_STAGES)
    for i, name in enumerate(PIPELINE_STAGES, start=1):
        print(f"Stage {i} of {n}: {name}", flush=True)
        try:
            run_stage(name)
        except Exception as exc:
            print("", flush=True)
            print("ERROR: Pipeline stopped because a stage failed.", flush=True)
            print(f"Failed stage: {i} of {n}, module src.{name}", flush=True)
            print(f"Exception type: {type(exc).__name__}", flush=True)
            print(f"Message: {exc}", flush=True)
            print("", flush=True)
            print("Full traceback:", flush=True)
            traceback.print_exc()
            return 1

    print("OK: All pipeline stages finished.", flush=True)

    haz_clean_hits: list[str] = []
    try:
        haz_clean_hits = _scan_outputs_for_legacy_haz_clean(config.OUTPUT_DIR)
        if haz_clean_hits:
            print(
                "ERROR: Legacy token Haz_clean still found under OUTPUT_DIR (csv/xlsx/md). "
                "Old field cleanup is incomplete.",
                flush=True,
            )
            for h in haz_clean_hits[:40]:
                print(f"  {h}", flush=True)
            if len(haz_clean_hits) > 40:
                print(f"  … and {len(haz_clean_hits) - 40} more hits.", flush=True)
    except Exception as exc:
        print(f"WARNING: Haz_clean output scan failed: {exc}", flush=True)

    if config.USE_REAL_ONLY:
        try:
            _append_quality_revision_run_log(haz_clean_hits=haz_clean_hits)
            print(f"OK: Appended quality revision block to {ROOT / 'RUN_LOG.md'}", flush=True)
        except Exception as exc:
            print(f"WARNING: Could not append RUN_LOG quality block: {exc}", flush=True)
        try:
            _append_local_chain_dedup_revision_run_log()
            print(f"OK: Appended local-chain table revision block to {ROOT / 'RUN_LOG.md'}", flush=True)
        except Exception as exc:
            print(f"WARNING: Could not append local-chain revision RUN_LOG block: {exc}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
